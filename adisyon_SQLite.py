import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta

class CafeAdisyonProgrami:
    @staticmethod
    def adapt_datetime(dt):
        return dt.strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def convert_datetime(text):
        return datetime.strptime(text.decode("utf-8"), "%Y-%m-%d %H:%M:%S")

    def __init__(self, root):
        self.root = root
        self.root.title("Kafe Adisyon Programı")
        self.root.geometry("1200x730")
        self.root.state('zoomed')
        
        # SQLite datetime dönüşümlerini kaydet
        sqlite3.register_adapter(datetime, self.adapt_datetime)
        sqlite3.register_converter("TIMESTAMP", self.convert_datetime)
        
        # Veritabanı bağlantısı
        self.conn = sqlite3.connect(
            'kafe.db',
            detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES
        )
        self.create_tables()
        
        # Arayüz bileşenlerini oluştur
        self.main_frame = None
        self.masalar_canvas = None
        self.masalar_frame = None
        self.masa_buttons = {}
        
        self.setup_ui()
        self.load_masalar()
        self.check_masa_durum()

    def setup_ui(self):
        """Arayüz bileşenlerini oluşturur"""
        # Ana frame
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Başlık
        tk.Label(
            self.main_frame, 
            text="MASA DURUMLARI", 
            font=('Arial', 20, 'bold'),
            pady=20
        ).pack(fill=tk.X)
        
        # Masalar için Canvas
        self.masalar_canvas = tk.Canvas(self.main_frame, highlightthickness=0)
        self.masalar_canvas.pack(fill=tk.BOTH, expand=True)
        
        # Masalar frame
        self.masalar_frame = tk.Frame(self.masalar_canvas)
        self.masalar_canvas.create_window((0, 0), window=self.masalar_frame, anchor="nw")
        
        # Yönetim butonları
        yonetim_frame = tk.LabelFrame(
            self.main_frame,
            text=" Yönetim Paneli ",
            font=('Arial', 10),
            bd=2,
            relief=tk.GROOVE,
            padx=10,
            pady=10
        )
        yonetim_frame.pack(fill=tk.X, pady=(10, 0))
        
        button_style = {'font': ('Arial', 10), 'width': 15, 'height': 1}
        
        tk.Button(
            yonetim_frame, 
            text="Ürün Yönetimi",
            relief=tk.GROOVE, 
            command=self.urun_yonetimi,
            **button_style
        ).pack(side=tk.LEFT, padx=10, pady=5)
        
        tk.Button(
            yonetim_frame, 
            text="Masa Yönetimi",
            relief=tk.GROOVE, 
            command=self.masa_yonetimi,
            **button_style
        ).pack(side=tk.LEFT, padx=10, pady=5)
        
        tk.Button(
            yonetim_frame, 
            text="Personel Yönetimi",
            relief=tk.GROOVE, 
            command=self.personel_yonetimi,
            **button_style
        ).pack(side=tk.LEFT, padx=10, pady=5)
        
        tk.Button(
            yonetim_frame, 
            text="Raporlar",
            relief=tk.GROOVE, 
            command=self.raporlari_ac,
            **button_style
        ).pack(side=tk.LEFT, padx=10, pady=5)
        
        # Yeniden boyutlandırma olayları
        self.masalar_frame.bind("<Configure>", lambda e: self.masalar_canvas.configure(scrollregion=self.masalar_frame.bbox("all")))
        self.masalar_canvas.bind("<Configure>", lambda e: self.masalar_canvas.itemconfig(1, width=e.width))

    def create_tables(self):
        """Veritabanı tablolarını oluşturur"""
        cursor = self.conn.cursor()

        cursor.execute('''
        CREATE TABLE IF NOT EXISTS masalar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            masa_no TEXT NOT NULL UNIQUE,
            durum TEXT DEFAULT 'Boş',
            son_islem_tarihi TIMESTAMP
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS kategoriler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ad TEXT NOT NULL UNIQUE
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS urunler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kategori_id INTEGER,
            ad TEXT NOT NULL,
            fiyat REAL NOT NULL,
            stok INTEGER DEFAULT 0,
            FOREIGN KEY (kategori_id) REFERENCES kategoriler(id)
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS personel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ad TEXT NOT NULL,
            pozisyon TEXT
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS siparisler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            masa_no TEXT NOT NULL,
            urun_id INTEGER NOT NULL,
            adet INTEGER NOT NULL,
            tarih TIMESTAMP NOT NULL,
            durum TEXT DEFAULT 'Aktif',
            FOREIGN KEY (urun_id) REFERENCES urunler(id)
        )''')
        
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS odemeler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            masa_no TEXT NOT NULL,
            tutar REAL NOT NULL,
            odeme_tipi TEXT NOT NULL,
            tarih TIMESTAMP NOT NULL
        )''')
        
        self.conn.commit()

    def load_masalar(self):
        """Masaları veritabanından yükler ve arayüzde gösterir"""
        if not hasattr(self, 'masalar_frame') or not self.masalar_frame:
            return
            
        # Önceki masaları temizle
        for widget in self.masalar_frame.winfo_children():
            widget.destroy()
        
        cursor = self.conn.cursor()
        cursor.execute("SELECT masa_no, durum, son_islem_tarihi FROM masalar ORDER BY CAST(masa_no AS INTEGER)")
        masalar = cursor.fetchall()
        
        self.masa_buttons = {}
        
        # Masaları 6 sütun halinde yerleştir
        for i, (masa_no, durum, son_islem) in enumerate(masalar):
            row = i // 6
            col = i % 6
            
            # Masa bakiyesini hesapla
            cursor.execute('''
            SELECT SUM(u.fiyat * s.adet) 
            FROM siparisler s 
            JOIN urunler u ON s.urun_id = u.id 
            WHERE s.masa_no=? AND s.durum='Aktif'
            ''', (masa_no,))
            bakiye = cursor.fetchone()[0] or 0.0
            
            btn_text = f"Masa {masa_no}\n{durum}\nBakiye: {bakiye:.2f} TL"
            
            # Renk belirleme
            if durum == "Boş":
                bg_color = "lightgreen"
            else:
                now = datetime.now()
                son_islem_time = son_islem if isinstance(son_islem, datetime) else (datetime.strptime(son_islem, "%Y-%m-%d %H:%M:%S") if son_islem else now)
                if (now - son_islem_time) > timedelta(minutes=30):
                    bg_color = "purple"
                else:
                    bg_color = "salmon"
            
            btn = tk.Button(
                self.masalar_frame,
                text=btn_text,
                width=18,
                height=8,
                relief=tk.RIDGE,  # 
                command=lambda m=masa_no: self.masa_penceresi_ac(m),
                bg=bg_color,
                font=('Arial', 10),
                padx=10,
                pady=10,
                wraplength=150
            )
            btn.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
            
            # Grid hücrelerinin genişlemesini sağla
            self.masalar_frame.grid_columnconfigure(col, weight=1)
            self.masalar_frame.grid_rowconfigure(row, weight=1)
            
            self.masa_buttons[masa_no] = btn

    def masa_penceresi_ac(self, masa_no):
        masa_pencere = tk.Toplevel(self.root)
        masa_pencere.title(f"Masa {masa_no} - Adisyon")
        masa_pencere.geometry("1000x700")
        
        # Masa bilgilerini yükle
        cursor = self.conn.cursor()
        cursor.execute("SELECT durum FROM masalar WHERE masa_no=?", (masa_no,))
        masa_durum = cursor.fetchone()[0]
        
        # Masa başlık
        tk.Label(
            masa_pencere, 
            text=f"MASA {masa_no}", 
            font=('Arial', 16, 'bold'),
            pady=10
        ).pack(fill=tk.X)
        
        # Durum bilgisi
        durum_label = tk.Label(
            masa_pencere, 
            text=f"Durum: {masa_durum}", 
            font=('Arial', 12)
        )
        durum_label.pack()
        
        # Ürün ekleme frame
        urun_frame = tk.LabelFrame(
            masa_pencere, 
            text="Ürün Ekle", 
            padx=5, 
            pady=5,
            font=('Arial', 10)
        )
        urun_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Kategori seçimi
        tk.Label(urun_frame, text="Kategori:").grid(row=0, column=0, padx=5, pady=5)
        kategori_combobox = ttk.Combobox(urun_frame, state="readonly", width=30)
        kategori_combobox.grid(row=0, column=1, padx=5, pady=5)
        
        # Ürün seçimi
        tk.Label(urun_frame, text="Ürün:").grid(row=1, column=0, padx=5, pady=5)
        urun_combobox = ttk.Combobox(urun_frame, state="readonly", width=30)
        urun_combobox.grid(row=1, column=1, padx=5, pady=5)
        
        # Adet
        tk.Label(urun_frame, text="Adet:").grid(row=2, column=0, padx=5, pady=5)
        adet_spinbox = tk.Spinbox(urun_frame, from_=1, to=20, width=5)
        adet_spinbox.grid(row=2, column=1, padx=5, pady=5, sticky=tk.W)
        
        # Kategorileri yükle
        cursor.execute("SELECT ad FROM kategoriler ORDER BY ad")
        kategoriler = [row[0] for row in cursor.fetchall()]
        kategori_combobox['values'] = kategoriler
        if kategoriler:
            kategori_combobox.current(0)
        
        # Kategori değişince ürünleri yükle
        def kategori_degisti(event):
            selected_kategori = kategori_combobox.get()
            cursor.execute("SELECT id FROM kategoriler WHERE ad=?", (selected_kategori,))
            kategori_id = cursor.fetchone()[0]
            
            cursor.execute("SELECT id, ad FROM urunler WHERE kategori_id=? ORDER BY ad", (kategori_id,))
            urunler = [f"{row[0]} - {row[1]}" for row in cursor.fetchall()]
            urun_combobox['values'] = urunler
            if urunler:
                urun_combobox.current(0)
        
        kategori_combobox.bind("<<ComboboxSelected>>", kategori_degisti)
        if kategoriler:
            kategori_degisti(None)
        
        # Sepet frame
        sepet_frame = tk.LabelFrame(
            masa_pencere, 
            text="Adisyon Detayı", 
            padx=5, 
            pady=5,
            font=('Arial', 10)
        )
        sepet_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Sepet treeview
        columns = ("Ürün", "Adet", "Birim Fiyat", "Toplam", "Tarih")
        sepet_tree = ttk.Treeview(
            sepet_frame, 
            columns=columns, 
            show="headings",
            selectmode="browse"
        )
        
        # Kolon genişlikleri
        sepet_tree.heading("Ürün", text="Ürün")
        sepet_tree.heading("Adet", text="Adet")
        sepet_tree.heading("Birim Fiyat", text="Birim Fiyat")
        sepet_tree.heading("Toplam", text="Toplam")
        sepet_tree.heading("Tarih", text="Tarih")
        
        sepet_tree.column("Ürün", width=200, anchor=tk.W)
        sepet_tree.column("Adet", width=50, anchor=tk.CENTER)
        sepet_tree.column("Birim Fiyat", width=100, anchor=tk.E)
        sepet_tree.column("Toplam", width=100, anchor=tk.E)
        sepet_tree.column("Tarih", width=150, anchor=tk.CENTER)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(sepet_frame, orient=tk.VERTICAL, command=sepet_tree.yview)
        sepet_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        sepet_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Toplam etiketi
        toplam_label = tk.Label(
            sepet_frame, 
            text="Toplam: 0.00 TL", 
            font=('Arial', 12, 'bold')
        )
        toplam_label.pack(side=tk.BOTTOM, fill=tk.X, pady=5)
        
        # Siparişleri yükleme fonksiyonu (eksik olan fonksiyon)
        def siparisleri_yukle():
            for item in sepet_tree.get_children():
                sepet_tree.delete(item)
            
            cursor.execute('''
            SELECT u.ad, s.adet, u.fiyat, (s.adet * u.fiyat) as toplam, s.tarih 
            FROM siparisler s 
            JOIN urunler u ON s.urun_id = u.id 
            WHERE s.masa_no=? AND s.durum='Aktif'
            ORDER BY s.tarih
            ''', (masa_no,))
            
            toplam = 0.0
            for urun_ad, adet, fiyat, urun_toplam, tarih in cursor.fetchall():
                formatted_tarih = tarih.strftime("%d.%m.%Y %H:%M")  # Direkt formatlama
                sepet_tree.insert("", tk.END, values=(
                    urun_ad, 
                    adet, 
                    f"{fiyat:.2f}", 
                    f"{urun_toplam:.2f}",
                    formatted_tarih
                ))
                toplam += urun_toplam
            
            toplam_label.config(text=f"Toplam: {toplam:.2f} TL")
            
            # Masa bakiyesini ana ekranda güncelle
            self.load_masalar()
        
        # Siparişleri ilk yükleme
        siparisleri_yukle()
    
        def adapt_datetime(self, dt):
            return dt.strftime("%Y-%m-%d %H:%M:%S")  # SQLite'ın anlayacağı formatta tarih

        # Sepete ekle butonu
        def sepete_ekle():
            urun = urun_combobox.get()
            if not urun:
                messagebox.showwarning("Uyarı", "Ürün seçmelisiniz!")
                return
            
            try:
                adet = int(adet_spinbox.get())
                if adet <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("Uyarı", "Geçerli bir adet giriniz!")
                return
            
            urun_id = int(urun.split(" - ")[0])
            
            # Ürün bilgilerini al
            cursor.execute("SELECT ad, fiyat FROM urunler WHERE id=?", (urun_id,))
            urun_ad, fiyat = cursor.fetchone()
            
            # Siparişi veritabanına ekle (Python 3.12+ uyumlu)
            tarih = datetime.now()
            cursor.execute('''
                INSERT INTO siparisler (masa_no, urun_id, adet, tarih)
                VALUES (?, ?, ?, ?)
            ''', (masa_no, urun_id, adet, tarih.strftime("%Y-%m-%d %H:%M:%S")))  # Direkt formatlama

            # Masayı dolu olarak işaretle ve son işlem tarihini güncelle
            cursor.execute('''
                UPDATE masalar 
                SET durum='Dolu', son_islem_tarihi=?
                WHERE masa_no=?
            ''', (tarih.strftime("%Y-%m-%d %H:%M:%S"), masa_no))  # Sadece 2 parametre: tarih ve masa_no
            
            self.conn.commit()
            
            # Durumu güncelle
            durum_label.config(text="Durum: Dolu")
            self.masa_buttons[masa_no].config(bg="salmon")
            
            # Sepeti yenile
            siparisleri_yukle()
            
            messagebox.showinfo("Başarılı", "Ürün başarıyla eklendi!")
        
        ekle_btn = tk.Button(
            urun_frame, 
            text="Ekle",
            relief=tk.GROOVE, 
            command=sepete_ekle,
            width=15
        )
        ekle_btn.grid(row=3, column=0, columnspan=2, pady=5, sticky=tk.EW)
        
        # Sipariş silme fonksiyonu
        def siparis_sil():
            selected = sepet_tree.selection()
            if not selected:
                messagebox.showwarning("Uyarı", "Silinecek ürünü seçmelisiniz!")
                return
            
            urun_ad = sepet_tree.item(selected[0])['values'][0]
            
            if messagebox.askyesno("Onay", f"{urun_ad} ürününü silmek istediğinize emin misiniz?"):
                cursor.execute('''
                DELETE FROM siparisler 
                WHERE masa_no=? AND urun_id=(
                    SELECT id FROM urunler WHERE ad=?
                ) AND durum='Aktif'
                ''', (masa_no, urun_ad))
                
                # Eğer başka sipariş kalmadıysa masayı boş yap
                cursor.execute("SELECT COUNT(*) FROM siparisler WHERE masa_no=? AND durum='Aktif'", (masa_no,))
                if cursor.fetchone()[0] == 0:
                    cursor.execute("UPDATE masalar SET durum='Boş' WHERE masa_no=?", (masa_no,))
                    durum_label.config(text="Durum: Boş")
                    self.masa_buttons[masa_no].config(bg="lightgreen")
                else:
                    # Son işlem tarihini güncelle
                    cursor.execute('''
                    UPDATE masalar 
                    SET son_islem_tarihi=?
                    WHERE masa_no=?
                    ''', (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), masa_no))
                
                self.conn.commit()
                siparisleri_yukle()
        
        # Adet güncelleme butonu
        def adet_guncelle():
            selected = sepet_tree.selection()
            if not selected:
                messagebox.showwarning("Uyarı", "Güncellenecek ürünü seçmelisiniz!")
                return
            
            urun_ad = sepet_tree.item(selected[0])['values'][0]
            eski_adet = sepet_tree.item(selected[0])['values'][1]
            
            # Yeni adet için input penceresi
            guncelle_pencere = tk.Toplevel(masa_pencere)
            guncelle_pencere.title("Adet Güncelle")
            
            tk.Label(guncelle_pencere, text=f"{urun_ad} ürünü için yeni adet:").pack(padx=10, pady=5)
            
            yeni_adet_spin = tk.Spinbox(guncelle_pencere, from_=1, to=20, width=5)
            yeni_adet_spin.pack(padx=10, pady=5)
            yeni_adet_spin.delete(0, tk.END)
            yeni_adet_spin.insert(0, eski_adet)
            
            def guncelle():
                try:
                    yeni_adet = int(yeni_adet_spin.get())
                    if yeni_adet <= 0:
                        raise ValueError
                except ValueError:
                    messagebox.showwarning("Uyarı", "Geçerli bir adet giriniz!")
                    return
                
                cursor.execute('''
                UPDATE siparisler 
                SET adet=?, tarih=?
                WHERE masa_no=? AND urun_id=(
                    SELECT id FROM urunler WHERE ad=?
                ) AND durum='Aktif'
                ''', (yeni_adet, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), masa_no, urun_ad))
                
                self.conn.commit()
                siparisleri_yukle()
                guncelle_pencere.destroy()
            
            tk.Button(guncelle_pencere, text="Güncelle", command=guncelle).pack(pady=5)
        
        guncelle_btn = tk.Button(
            sepet_frame, 
            text="Adet Güncelle",
            relief=tk.GROOVE, 
            command=adet_guncelle,
            width=15
        )
        guncelle_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        # İndirim butonu - DÜZELTİLMİŞ VERSİYON
        def indirim_yap():
            selected = sepet_tree.selection()
            if not selected:
                messagebox.showwarning("Uyarı", "İndirim yapılacak ürünü seçmelisiniz!")
                return
            
            urun_bilgisi = sepet_tree.item(selected[0])['values']
            urun_ad = urun_bilgisi[0]
            eski_adet = urun_bilgisi[1]
            eski_fiyat = float(urun_bilgisi[2])
            
            # İndirim yüzdesi için input penceresi
            indirim_pencere = tk.Toplevel(masa_pencere)
            indirim_pencere.title("İndirim Yap")
            
            tk.Label(indirim_pencere, text=f"{urun_ad} ürününe uygulanacak indirim (%):").pack(padx=10, pady=5)
            
            indirim_spin = tk.Spinbox(indirim_pencere, from_=0, to=100, width=5)
            indirim_spin.pack(padx=10, pady=5)
            indirim_spin.delete(0, tk.END)
            indirim_spin.insert(0, "10")
            
            def indirim_uygula():
                try:
                    indirim = float(indirim_spin.get())
                    if indirim < 0 or indirim > 100:
                        raise ValueError
                except ValueError:
                    messagebox.showwarning("Uyarı", "Geçerli bir indirim yüzdesi giriniz!")
                    return
                
                # Mevcut ürünün ID'sini al
                cursor.execute("SELECT id FROM urunler WHERE ad=?", (urun_ad,))
                urun_id = cursor.fetchone()[0]
                
                # İndirimli fiyatı hesapla
                indirimli_fiyat = eski_fiyat * (100 - indirim) / 100
                
                # Siparişi güncelle (yeni fiyatla)
                cursor.execute('''
                UPDATE siparisler 
                SET urun_id=?, tarih=?
                WHERE masa_no=? AND urun_id=? AND durum='Aktif'
                ''', (urun_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), masa_no, urun_id))
                
                # Masanın son işlem tarihini güncelle
                cursor.execute('''
                UPDATE masalar 
                SET son_islem_tarihi=?
                WHERE masa_no=?
                ''', (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), masa_no))
                
                self.conn.commit()
                siparisleri_yukle()
                indirim_pencere.destroy()
                messagebox.showinfo("Başarılı", "İndirim başarıyla uygulandı!")
            
            tk.Button(indirim_pencere, text="İndirim Uygula", relief=tk.GROOVE, command=indirim_uygula).pack(pady=5)
        
        indirim_btn = tk.Button(
            sepet_frame, 
            text="İndirim Yap",
            relief=tk.GROOVE, 
            command=indirim_yap,
            width=15
        )
        indirim_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Hesap kapatma butonu
        def hesap_kapat():
            cursor.execute('''
            SELECT SUM(u.fiyat * s.adet) 
            FROM siparisler s 
            JOIN urunler u ON s.urun_id = u.id 
            WHERE s.masa_no=? AND s.durum='Aktif'
            ''', (masa_no,))
            
            toplam_tutar = cursor.fetchone()[0] or 0.0
            
            if toplam_tutar <= 0:
                messagebox.showwarning("Uyarı", "Kapatılacak hesap bulunamadı!")
                return
            
            # Ödeme penceresi
            odeme_pencere = tk.Toplevel(masa_pencere)
            odeme_pencere.title("Hesap Kapat")
            
            tk.Label(
                odeme_pencere, 
                text=f"Toplam Tutar: {toplam_tutar:.2f} TL", 
                font=('Arial', 12, 'bold')
            ).pack(pady=10)
            
            tk.Label(odeme_pencere, text="Ödeme Tipi:").pack()
            odeme_tipi = ttk.Combobox(
                odeme_pencere, 
                values=["Nakit", "Kredi Kartı", "Havale"], 
                state="readonly"
            )
            odeme_tipi.pack(pady=5)
            odeme_tipi.current(0)
            
            def odemeyi_tamamla():
                # Siparişleri tamamlandı olarak işaretle
                cursor.execute("UPDATE siparisler SET durum='Tamamlandı' WHERE masa_no=? AND durum='Aktif'", (masa_no,))
                
                # Ödeme kaydını ekle
                tarih = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute('''
                INSERT INTO odemeler (masa_no, tutar, odeme_tipi, tarih)
                VALUES (?, ?, ?, ?)
                ''', (masa_no, toplam_tutar, odeme_tipi.get(), tarih))
                
                # Masayı boş yap
                cursor.execute("UPDATE masalar SET durum='Boş' WHERE masa_no=?", (masa_no,))
                
                self.conn.commit()
                
                # Kapat
                messagebox.showinfo("Başarılı", "Hesap başarıyla kapatıldı!")
                odeme_pencere.destroy()
                masa_pencere.destroy()
                
                # Ana ekranı güncelle
                self.load_masalar()
            
            tk.Button(
                odeme_pencere, 
                text="Ödemeyi Tamamla",
                relief=tk.GROOVE, 
                command=odemeyi_tamamla
            ).pack(pady=10)
        
        kapat_btn = tk.Button(
            sepet_frame, 
            text="Hesap Kapat",
            relief=tk.GROOVE, 
            command=hesap_kapat,
            width=15
        )
        kapat_btn.pack(side=tk.RIGHT, padx=5, pady=5)

        kapat_btn.pack(side=tk.RIGHT, padx=5, pady=5)
    
    def urun_yonetimi(self):
        yonetim_pencere = tk.Toplevel(self.root)
        yonetim_pencere.title("Ürün Yönetimi")
        yonetim_pencere.geometry("800x600")
        
        # Kategori yönetimi frame
        kategori_frame = tk.LabelFrame(yonetim_pencere, text="Kategori Yönetimi", padx=5, pady=5)
        kategori_frame.pack(fill=tk.X, padx=10, pady=5)
        
        tk.Label(kategori_frame, text="Kategori Adı:").grid(row=0, column=0, padx=5, pady=5)
        kategori_ad_entry = tk.Entry(kategori_frame, width=30)
        kategori_ad_entry.grid(row=0, column=1, padx=5, pady=5)
        
        def kategori_ekle():
            kategori_ad = kategori_ad_entry.get().strip()
            if not kategori_ad:
                messagebox.showwarning("Uyarı", "Kategori adı boş olamaz!")
                return
            
            cursor = self.conn.cursor()
            try:
                cursor.execute("INSERT INTO kategoriler (ad) VALUES (?)", (kategori_ad,))
                self.conn.commit()
                messagebox.showinfo("Başarılı", "Kategori başarıyla eklendi!")
                kategori_ad_entry.delete(0, tk.END)
            except sqlite3.IntegrityError:
                messagebox.showerror("Hata", "Bu kategori zaten var!")
        
        tk.Button(kategori_frame, text="Kategori Ekle", relief=tk.GROOVE, command=kategori_ekle).grid(row=0, column=2, padx=5, pady=5)
        
        # Ürün yönetimi frame
        urun_frame = tk.LabelFrame(yonetim_pencere, text="Ürün Yönetimi", padx=5, pady=5)
        urun_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Ürün listesi
        columns = ("ID", "Kategori", "Ürün Adı", "Fiyat", "Stok")
        urun_tree = ttk.Treeview(urun_frame, columns=columns, show="headings")
        
        for col in columns:
            urun_tree.heading(col, text=col)
            urun_tree.column(col, width=100, anchor=tk.CENTER)
        
        urun_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        # Ürün formu
        form_frame = tk.Frame(urun_frame)
        form_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)
        
        tk.Label(form_frame, text="Kategori:").grid(row=0, column=0, padx=5, pady=5)
        kategori_combobox = ttk.Combobox(form_frame, width=27, state="readonly")
        kategori_combobox.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(form_frame, text="Ürün Adı:").grid(row=1, column=0, padx=5, pady=5)
        urun_ad_entry = tk.Entry(form_frame, width=30)
        urun_ad_entry.grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(form_frame, text="Fiyat:").grid(row=2, column=0, padx=5, pady=5)
        urun_fiyat_entry = tk.Entry(form_frame, width=30)
        urun_fiyat_entry.grid(row=2, column=1, padx=5, pady=5)
        
        tk.Label(form_frame, text="Stok:").grid(row=3, column=0, padx=5, pady=5)
        urun_stok_entry = tk.Entry(form_frame, width=30)
        urun_stok_entry.grid(row=3, column=1, padx=5, pady=5)
        
        # Butonlar
        btn_frame = tk.Frame(form_frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        
        tk.Button(btn_frame, text="Ekle", relief=tk.GROOVE, command=lambda: self.urun_ekle(
            kategori_combobox.get(),
            urun_ad_entry.get(),
            urun_fiyat_entry.get(),
            urun_stok_entry.get(),
            urun_tree
        )).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Güncelle", relief=tk.GROOVE, command=lambda: self.urun_guncelle(
            urun_tree,
            kategori_combobox.get(),
            urun_ad_entry.get(),
            urun_fiyat_entry.get(),
            urun_stok_entry.get()
        )).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Sil", relief=tk.GROOVE, command=lambda: self.urun_sil(urun_tree)).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Temizle", relief=tk.GROOVE, command=lambda: self.urun_form_temizle(
            kategori_combobox,
            urun_ad_entry,
            urun_fiyat_entry,
            urun_stok_entry,
            urun_tree
        )).pack(side=tk.LEFT, padx=5)
        
        # Kategorileri yükle
        cursor = self.conn.cursor()
        cursor.execute("SELECT ad FROM kategoriler ORDER BY ad")
        kategoriler = [row[0] for row in cursor.fetchall()]
        kategori_combobox['values'] = kategoriler
        if kategoriler:
            kategori_combobox.current(0)
        
        # Ürünleri yükle
        self.urunleri_yukle(urun_tree)
        
        def kategori_sil(self):
            """ Kategori silme penceresini açar """
            # Pencere oluştur
            sil_penceresi = tk.Toplevel()
            sil_penceresi.title("Kategori Sil")
            sil_penceresi.geometry("300x200")

            # Kategori listesi
            tk.Label(sil_penceresi, text="Silinecek Kategori:").pack(pady=5)
            
            self.lb_kategoriler = tk.Listbox(sil_penceresi)
            self.lb_kategoriler.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            # Mevcut kategorileri yükle
            self.cursor.execute("SELECT ad FROM kategoriler")
            for kategori in self.cursor.fetchall():
                self.lb_kategoriler.insert(tk.END, kategori[0])

            # Sil butonu
            tk.Button(sil_penceresi, text="SİL", command=self._kategori_sil_onayla, 
                    bg="red", fg="white").pack(pady=10)

        def _kategori_sil_onayla(self):
            """ Seçili kategoriyi siler (öncesinde onay ister) """
            secim = self.lb_kategoriler.curselection()
            if not secim:
                messagebox.showwarning("Uyarı", "Lütfen bir kategori seçin!")
                return
            
            kategori_ad = self.lb_kategoriler.get(secim[0])
            
            # Kategoriye bağlı ürün var mı kontrol et
            self.cursor.execute("SELECT COUNT(*) FROM urunler WHERE kategori_id=(SELECT id FROM kategoriler WHERE ad=?)", (kategori_ad,))
            urun_sayisi = self.cursor.fetchone()[0]
            
            if urun_sayisi > 0:
                messagebox.showerror("Hata", f"Bu kategoride {urun_sayisi} ürün var! Önce ürünleri silin.")
                return
            
            # Onay iste
            if not messagebox.askyesno("Onay", f"'{kategori_ad}' kategorisini silmek istediğinize emin misiniz?"):
                return
            
            # Silme işlemi
            try:
                self.cursor.execute("DELETE FROM kategoriler WHERE ad=?", (kategori_ad,))
                self.conn.commit()
                messagebox.showinfo("Başarılı", "Kategori silindi!")
                self.lb_kategoriler.delete(secim[0])  # Listeden kaldır
            except Exception as e:
                messagebox.showerror("Hata", f"Silme başarısız: {str(e)}")

        # Treeview seçim olayı
        def urun_sec(event):
            selected = urun_tree.selection()
            if not selected:
                return
            
            item = urun_tree.item(selected[0])
            values = item['values']
            
            kategori_combobox.set(values[1])
            urun_ad_entry.delete(0, tk.END)
            urun_ad_entry.insert(0, values[2])
            urun_fiyat_entry.delete(0, tk.END)
            urun_fiyat_entry.insert(0, str(values[3]))
            urun_stok_entry.delete(0, tk.END)
            urun_stok_entry.insert(0, str(values[4]))
        
        urun_tree.bind("<<TreeviewSelect>>", urun_sec)
    
    def urunleri_yukle(self, treeview):
        for item in treeview.get_children():
            treeview.delete(item)
        
        cursor = self.conn.cursor()
        cursor.execute('''
        SELECT u.id, k.ad, u.ad, u.fiyat, u.stok 
        FROM urunler u 
        LEFT JOIN kategoriler k ON u.kategori_id = k.id
        ORDER BY k.ad, u.ad
        ''')
        
        for row in cursor.fetchall():
            treeview.insert("", tk.END, values=row)
    
    def urun_ekle(self, kategori, ad, fiyat, stok, treeview):
        if not all([kategori, ad, fiyat]):
            messagebox.showwarning("Uyarı", "Kategori, ürün adı ve fiyat zorunludur!")
            return
        
        try:
            fiyat = float(fiyat)
            if fiyat <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Uyarı", "Geçerli bir fiyat giriniz!")
            return
        
        try:
            stok = int(stok) if stok else 0
        except ValueError:
            messagebox.showwarning("Uyarı", "Geçerli bir stok miktarı giriniz!")
            return
        
        cursor = self.conn.cursor()
        
        # Kategori ID'sini al
        cursor.execute("SELECT id FROM kategoriler WHERE ad=?", (kategori,))
        kategori_id = cursor.fetchone()
        
        if not kategori_id:
            messagebox.showwarning("Uyarı", "Geçerli bir kategori seçiniz!")
            return
        
        kategori_id = kategori_id[0]
        
        try:
            cursor.execute('''
            INSERT INTO urunler (kategori_id, ad, fiyat, stok)
            VALUES (?, ?, ?, ?)
            ''', (kategori_id, ad, fiyat, stok))
            
            self.conn.commit()
            messagebox.showinfo("Başarılı", "Ürün başarıyla eklendi!")
            self.urunleri_yukle(treeview)
        except sqlite3.IntegrityError:
            messagebox.showerror("Hata", "Bu ürün adı zaten var!")
    
    def urun_guncelle(self, treeview, kategori, ad, fiyat, stok):
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Güncellenecek ürünü seçmelisiniz!")
            return
        
        urun_id = treeview.item(selected[0])['values'][0]
        
        if not all([kategori, ad, fiyat]):
            messagebox.showwarning("Uyarı", "Kategori, ürün adı ve fiyat zorunludur!")
            return
        
        try:
            fiyat = float(fiyat)
            if fiyat <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Uyarı", "Geçerli bir fiyat giriniz!")
            return
        
        try:
            stok = int(stok) if stok else 0
        except ValueError:
            messagebox.showwarning("Uyarı", "Geçerli bir stok miktarı giriniz!")
            return
        
        cursor = self.conn.cursor()
        
        # Kategori ID'sini al
        cursor.execute("SELECT id FROM kategoriler WHERE ad=?", (kategori,))
        kategori_id = cursor.fetchone()
        
        if not kategori_id:
            messagebox.showwarning("Uyarı", "Geçerli bir kategori seçiniz!")
            return
        
        kategori_id = kategori_id[0]
        
        try:
            cursor.execute('''
            UPDATE urunler 
            SET kategori_id=?, ad=?, fiyat=?, stok=?
            WHERE id=?
            ''', (kategori_id, ad, fiyat, stok, urun_id))
            
            self.conn.commit()
            messagebox.showinfo("Başarılı", "Ürün başarıyla güncellendi!")
            self.urunleri_yukle(treeview)
        except sqlite3.IntegrityError:
            messagebox.showerror("Hata", "Bu ürün adı zaten var!")
    
    def urun_sil(self, treeview):
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Silinecek ürünü seçmelisiniz!")
            return
        
        urun_id = treeview.item(selected[0])['values'][0]
        
        if messagebox.askyesno("Onay", "Bu ürünü silmek istediğinize emin misiniz?"):
            cursor = self.conn.cursor()
            
            try:
                # Ürünün siparişlerde kullanılıp kullanılmadığını kontrol et
                cursor.execute("SELECT COUNT(*) FROM siparisler WHERE urun_id=?", (urun_id,))
                count = cursor.fetchone()[0]
                
                if count > 0:
                    messagebox.showwarning("Uyarı", "Bu ürün siparişlerde kullanılmış, silinemez!")
                    return
                
                cursor.execute("DELETE FROM urunler WHERE id=?", (urun_id,))
                self.conn.commit()
                messagebox.showinfo("Başarılı", "Ürün başarıyla silindi!")
                self.urunleri_yukle(treeview)
            except Exception as e:
                messagebox.showerror("Hata", f"Ürün silinirken hata oluştu: {str(e)}")
    
    def urun_form_temizle(self, kategori_combobox, ad_entry, fiyat_entry, stok_entry, treeview):
        kategori_combobox.set('')
        ad_entry.delete(0, tk.END)
        fiyat_entry.delete(0, tk.END)
        stok_entry.delete(0, tk.END)
        treeview.selection_remove(treeview.selection())
    
    def masa_yonetimi(self):
        yonetim_pencere = tk.Toplevel(self.root)
        yonetim_pencere.title("Masa Yönetimi")
        yonetim_pencere.geometry("600x400")
        
        # Masa listesi
        columns = ("Masa No", "Durum")
        masa_tree = ttk.Treeview(yonetim_pencere, columns=columns, show="headings")
        
        for col in columns:
            masa_tree.heading(col, text=col)
            masa_tree.column(col, width=100, anchor=tk.CENTER)
        
        masa_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Masa formu
        form_frame = tk.Frame(yonetim_pencere)
        form_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)
        
        tk.Label(form_frame, text="Masa No:").grid(row=0, column=0, padx=5, pady=5)
        masa_no_entry = tk.Entry(form_frame, width=23)
        masa_no_entry.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(form_frame, text="Durum:").grid(row=1, column=0, padx=5, pady=5)
        masa_durum_combobox = ttk.Combobox(form_frame, values=["Boş", "Dolu"], state="readonly")
        masa_durum_combobox.grid(row=1, column=1, padx=5, pady=5)
        masa_durum_combobox.current(0)
        
        # Butonlar
        btn_frame = tk.Frame(form_frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        tk.Button(btn_frame, text="Ekle", relief=tk.GROOVE, command=lambda: self.masa_ekle(
            masa_no_entry.get(),
            masa_durum_combobox.get(),
            masa_tree
        )).pack(side=tk.LEFT, padx=10)
        
        tk.Button(btn_frame, text="Güncelle", relief=tk.GROOVE, command=lambda: self.masa_guncelle(
            masa_tree,
            masa_no_entry.get(),
            masa_durum_combobox.get()
        )).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Sil", relief=tk.GROOVE, command=lambda: self.masa_sil(masa_tree)).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Temizle", relief=tk.GROOVE, command=lambda: self.masa_form_temizle(
            masa_no_entry,
            masa_durum_combobox,
            masa_tree
        )).pack(side=tk.LEFT, padx=5)
        
        # Masaları yükle
        self.masalari_yukle(masa_tree)
        
        # Treeview seçim olayı
        def masa_sec(event):
            selected = masa_tree.selection()
            if not selected:
                return
            
            item = masa_tree.item(selected[0])
            values = item['values']
            
            masa_no_entry.delete(0, tk.END)
            masa_no_entry.insert(0, values[0])
            masa_durum_combobox.set(values[1])
        
        masa_tree.bind("<<TreeviewSelect>>", masa_sec)
    
    def masalari_yukle(self, treeview):
        for item in treeview.get_children():
            treeview.delete(item)
        
        cursor = self.conn.cursor()
        cursor.execute("SELECT masa_no, durum FROM masalar ORDER BY CAST(masa_no AS INTEGER)")
        
        for row in cursor.fetchall():
            treeview.insert("", tk.END, values=row)
    
    def masa_ekle(self, masa_no, durum, treeview):
        if not masa_no:
            messagebox.showwarning("Uyarı", "Masa numarası boş olamaz!")
            return
        
        cursor = self.conn.cursor()
        try:
            cursor.execute("INSERT INTO masalar (masa_no, durum) VALUES (?, ?)", (masa_no, durum))
            self.conn.commit()
            messagebox.showinfo("Başarılı", "Masa başarıyla eklendi!")
            self.masalari_yukle(treeview)
            self.load_masalar()
        except sqlite3.IntegrityError:
            messagebox.showerror("Hata", "Bu masa numarası zaten var!")
    
    def masa_guncelle(self, treeview, masa_no, durum):
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Güncellenecek masayı seçmelisiniz!")
            return
        
        eski_masa_no = treeview.item(selected[0])['values'][0]
        
        if not masa_no:
            messagebox.showwarning("Uyarı", "Masa numarası boş olamaz!")
            return
        
        cursor = self.conn.cursor()
        try:
            if eski_masa_no != masa_no:
                cursor.execute("UPDATE masalar SET masa_no=?, durum=? WHERE masa_no=?", 
                             (masa_no, durum, eski_masa_no))
            else:
                cursor.execute("UPDATE masalar SET durum=? WHERE masa_no=?", (durum, masa_no))
            
            self.conn.commit()
            messagebox.showinfo("Başarılı", "Masa başarıyla güncellendi!")
            self.masalari_yukle(treeview)
            self.load_masalar()
        except sqlite3.IntegrityError:
            messagebox.showerror("Hata", "Bu masa numarası zaten var!")
    
    def masa_sil(self, treeview):
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Silinecek masayı seçmelisiniz!")
            return
        
        masa_no = treeview.item(selected[0])['values'][0]
        
        if messagebox.askyesno("Onay", "Bu masayı silmek istediğinize emin misiniz?"):
            cursor = self.conn.cursor()
            
            try:
                # Masanın siparişlerde kullanılıp kullanılmadığını kontrol et
                cursor.execute("SELECT COUNT(*) FROM siparisler WHERE masa_no=?", (masa_no,))
                count = cursor.fetchone()[0]
                
                if count > 0:
                    messagebox.showwarning("Uyarı", "Bu masa siparişlerde kullanılmış, silinemez!")
                    return
                
                cursor.execute("DELETE FROM masalar WHERE masa_no=?", (masa_no,))
                self.conn.commit()
                messagebox.showinfo("Başarılı", "Masa başarıyla silindi!")
                self.masalari_yukle(treeview)
                self.load_masalar()
            except Exception as e:
                messagebox.showerror("Hata", f"Masa silinirken hata oluştu: {str(e)}")
    
    def masa_form_temizle(self, masa_no_entry, durum_combobox, treeview):
        masa_no_entry.delete(0, tk.END)
        durum_combobox.current(0)
        treeview.selection_remove(treeview.selection())
    
    def personel_yonetimi(self):
        yonetim_pencere = tk.Toplevel(self.root)
        yonetim_pencere.title("Personel Yönetimi")
        yonetim_pencere.geometry("600x400")
        
        # Personel listesi
        columns = ("ID", "Ad", "Pozisyon")
        personel_tree = ttk.Treeview(yonetim_pencere, columns=columns, show="headings")
        
        for col in columns:
            personel_tree.heading(col, text=col)
            personel_tree.column(col, width=100, anchor=tk.CENTER)
        
        personel_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Personel formu
        form_frame = tk.Frame(yonetim_pencere)
        form_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)
        
        tk.Label(form_frame, text="Ad:").grid(row=0, column=0, padx=5, pady=5)
        personel_ad_entry = tk.Entry(form_frame, width=20)
        personel_ad_entry.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(form_frame, text="Pozisyon:").grid(row=1, column=0, padx=5, pady=5)
        personel_pozisyon_entry = tk.Entry(form_frame, width=20)
        personel_pozisyon_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # Butonlar
        btn_frame = tk.Frame(form_frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        tk.Button(btn_frame, text="Ekle", command=lambda: self.personel_ekle(
            personel_ad_entry.get(),
            personel_pozisyon_entry.get(),
            personel_tree
        )).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Güncelle", command=lambda: self.personel_guncelle(
            personel_tree,
            personel_ad_entry.get(),
            personel_pozisyon_entry.get()
        )).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Sil", command=lambda: self.personel_sil(personel_tree)).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Temizle", command=lambda: self.personel_form_temizle(
            personel_ad_entry,
            personel_pozisyon_entry,
            personel_tree
        )).pack(side=tk.LEFT, padx=5)
        
        # Personelleri yükle
        self.personelleri_yukle(personel_tree)
        
        # Treeview seçim olayı
        def personel_sec(event):
            selected = personel_tree.selection()
            if not selected:
                return
            
            item = personel_tree.item(selected[0])
            values = item['values']
            
            personel_ad_entry.delete(0, tk.END)
            personel_ad_entry.insert(0, values[1])
            personel_pozisyon_entry.delete(0, tk.END)
            personel_pozisyon_entry.insert(0, values[2] if len(values) > 2 else "")
        
        personel_tree.bind("<<TreeviewSelect>>", personel_sec)
    
    def personelleri_yukle(self, treeview):
        for item in treeview.get_children():
            treeview.delete(item)
        
        cursor = self.conn.cursor()
        cursor.execute("SELECT id, ad, pozisyon FROM personel ORDER BY ad")
        
        for row in cursor.fetchall():
            treeview.insert("", tk.END, values=row)
    
    def personel_ekle(self, ad, pozisyon, treeview):
        if not ad:
            messagebox.showwarning("Uyarı", "Personel adı boş olamaz!")
            return
        
        cursor = self.conn.cursor()
        try:
            cursor.execute("INSERT INTO personel (ad, pozisyon) VALUES (?, ?)", (ad, pozisyon))
            self.conn.commit()
            messagebox.showinfo("Başarılı", "Personel başarıyla eklendi!")
            self.personelleri_yukle(treeview)
        except sqlite3.IntegrityError:
            messagebox.showerror("Hata", "Bu personel adı zaten var!")
    
    def personel_guncelle(self, treeview, ad, pozisyon):
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Güncellenecek personeli seçmelisiniz!")
            return
        
        personel_id = treeview.item(selected[0])['values'][0]
        
        if not ad:
            messagebox.showwarning("Uyarı", "Personel adı boş olamaz!")
            return
        
        cursor = self.conn.cursor()
        try:
            cursor.execute("UPDATE personel SET ad=?, pozisyon=? WHERE id=?", (ad, pozisyon, personel_id))
            self.conn.commit()
            messagebox.showinfo("Başarılı", "Personel başarıyla güncellendi!")
            self.personelleri_yukle(treeview)
        except sqlite3.IntegrityError:
            messagebox.showerror("Hata", "Bu personel adı zaten var!")
    
    def personel_sil(self, treeview):
        selected = treeview.selection()
        if not selected:
            messagebox.showwarning("Uyarı", "Silinecek personeli seçmelisiniz!")
            return
        
        personel_id = treeview.item(selected[0])['values'][0]
        
        if messagebox.askyesno("Onay", "Bu personeli silmek istediğinize emin misiniz?"):
            cursor = self.conn.cursor()
            
            try:
                # Personelin siparişlerde kullanılıp kullanılmadığını kontrol et
                cursor.execute("SELECT COUNT(*) FROM siparisler WHERE personel_id=?", (personel_id,))
                count = cursor.fetchone()[0]
                
                if count > 0:
                    messagebox.showwarning("Uyarı", "Bu personel siparişlerde kullanılmış, silinemez!")
                    return
                
                cursor.execute("DELETE FROM personel WHERE id=?", (personel_id,))
                self.conn.commit()
                messagebox.showinfo("Başarılı", "Personel başarıyla silindi!")
                self.personelleri_yukle(treeview)
            except Exception as e:
                messagebox.showerror("Hata", f"Personel silinirken hata oluştu: {str(e)}")
    
    def personel_form_temizle(self, ad_entry, pozisyon_entry, treeview):
        ad_entry.delete(0, tk.END)
        pozisyon_entry.delete(0, tk.END)
        treeview.selection_remove(treeview.selection())
    
    def raporlari_ac(self):
        rapor_pencere = tk.Toplevel(self.root)
        rapor_pencere.title("Raporlar")
        rapor_pencere.geometry("800x600")
        
        # Rapor seçimi
        rapor_frame = tk.LabelFrame(rapor_pencere, text="Rapor Seçimi", padx=0, pady=5)
        rapor_frame.pack(fill=tk.X, padx=10, pady=5)
        
        tk.Label(rapor_frame, text="Rapor Türü:").grid(row=0, column=0, padx=0, pady=5)
        rapor_combobox = ttk.Combobox(rapor_frame, values=[
            "Günlük Satış Raporu",
            "Aylık Satış Raporu",
            "Personel Performans Raporu",
            "Ürün Satış Raporu"
        ], state="readonly")
        rapor_combobox.grid(row=0, column=1, padx=5, pady=5)
        rapor_combobox.current(0)
        
        # Tarih aralığı
        tarih_frame = tk.Frame(rapor_frame)
        tarih_frame.grid(row=1, column=0, columnspan=2, pady=5)
        
        tk.Label(tarih_frame, text="Başlangıç:").pack(side=tk.LEFT, padx=5)
        baslangic_tarih_entry = tk.Entry(tarih_frame, width=10)
        baslangic_tarih_entry.pack(side=tk.LEFT, padx=5)
        baslangic_tarih_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        tk.Label(tarih_frame, text="Bitiş:").pack(side=tk.LEFT, padx=5)
        bitis_tarih_entry = tk.Entry(tarih_frame, width=10)
        bitis_tarih_entry.pack(side=tk.LEFT, padx=5)
        bitis_tarih_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # Rapor butonu
        tk.Button(rapor_frame, text="Rapor Oluştur", relief=tk.GROOVE, command=lambda: self.rapor_olustur(
            rapor_combobox.get(),
            baslangic_tarih_entry.get(),
            bitis_tarih_entry.get(),
            rapor_tree,
            toplam_label
        )).grid(row=0, column=2, columnspan=2, pady=5)
        
        # Rapor sonuçları
        rapor_sonuc_frame = tk.LabelFrame(rapor_pencere, text="Rapor Sonuçları", padx=5, pady=5)
        rapor_sonuc_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Rapor treeview
        rapor_tree = ttk.Treeview(rapor_sonuc_frame)
        rapor_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Toplam etiketi
        toplam_label = tk.Label(rapor_sonuc_frame, text="Toplam: 0.00 TL", font=('Arial', 12, 'bold'))
        toplam_label.pack(side=tk.BOTTOM, fill=tk.X, pady=5)
        
        # Excel'e aktar butonu
        tk.Button(rapor_sonuc_frame, text="Excel'e Aktar", relief=tk.GROOVE, command=lambda: self.raporu_excele_aktar(rapor_tree)).pack(side=tk.RIGHT, pady=5)
    
    def rapor_olustur(self, rapor_turu, baslangic_tarih, bitis_tarih, treeview, toplam_label):
        try:
            # Tarih formatını kontrol et
            datetime.strptime(baslangic_tarih, "%Y-%m-%d")
            datetime.strptime(bitis_tarih, "%Y-%m-%d")
        except ValueError:
            messagebox.showwarning("Uyarı", "Geçerli bir tarih formatı giriniz (YYYY-AA-GG)")
            return
        
        cursor = self.conn.cursor()
        
        # Treeview'ı temizle
        for item in treeview.get_children():
            treeview.delete(item)
        
        # Sütunları ayarla
        columns = []
        query = ""
        
        if rapor_turu == "Günlük Satış Raporu":
            columns = ["Tarih", "Toplam Satış", "Sipariş Sayısı"]
            query = '''
            SELECT 
                date(s.tarih) as Tarih,
                SUM(u.fiyat * s.adet) as ToplamSatis,
                COUNT(DISTINCT s.masa_no) as SiparisSayisi
            FROM siparisler s
            JOIN urunler u ON s.urun_id = u.id
            WHERE date(s.tarih) BETWEEN ? AND ?
            GROUP BY date(s.tarih)
            ORDER BY date(s.tarih)
            '''
        elif rapor_turu == "Aylık Satış Raporu":
            columns = ["Ay", "Toplam Satış", "Sipariş Sayısı"]
            query = '''
            SELECT 
                strftime('%Y-%m', s.tarih) as Ay,
                SUM(u.fiyat * s.adet) as ToplamSatis,
                COUNT(DISTINCT s.masa_no) as SiparisSayisi
            FROM siparisler s
            JOIN urunler u ON s.urun_id = u.id
            WHERE date(s.tarih) BETWEEN ? AND ?
            GROUP BY strftime('%Y-%m', s.tarih)
            ORDER BY strftime('%Y-%m', s.tarih)
            '''
        elif rapor_turu == "Personel Performans Raporu":
            columns = ["Personel", "Toplam Satış", "Sipariş Sayısı"]
            query = '''
            SELECT 
                p.ad as Personel,
                SUM(u.fiyat * s.adet) as ToplamSatis,
                COUNT(DISTINCT s.masa_no) as SiparisSayisi
            FROM siparisler s
            JOIN urunler u ON s.urun_id = u.id
            JOIN personel p ON s.personel_id = p.id
            WHERE date(s.tarih) BETWEEN ? AND ?
            GROUP BY p.ad
            ORDER BY SUM(u.fiyat * s.adet) DESC
            '''
        elif rapor_turu == "Ürün Satış Raporu":
            columns = ["Ürün", "Toplam Satış", "Satılan Adet"]
            query = '''
            SELECT 
                u.ad as Urun,
                SUM(u.fiyat * s.adet) as ToplamSatis,
                SUM(s.adet) as SatilanAdet
            FROM siparisler s
            JOIN urunler u ON s.urun_id = u.id
            WHERE date(s.tarih) BETWEEN ? AND ?
            GROUP BY u.ad
            ORDER BY SUM(u.fiyat * s.adet) DESC
            '''
        
        # Treeview sütunlarını ayarla
        treeview['columns'] = columns
        for col in columns:
            treeview.heading(col, text=col)
            treeview.column(col, width=100, anchor=tk.CENTER)
        
        # Raporu çalıştır
        cursor.execute(query, (baslangic_tarih, bitis_tarih))
        rapor_verileri = cursor.fetchall()
        
        # Toplamı hesapla
        toplam = 0.0
        if rapor_turu in ["Günlük Satış Raporu", "Aylık Satış Raporu", "Personel Performans Raporu", "Ürün Satış Raporu"]:
            toplam = sum(row[1] for row in rapor_verileri)
        
        # Verileri treeview'a ekle
        for row in rapor_verileri:
            treeview.insert("", tk.END, values=row)
        
        # Toplamı göster
        toplam_label.config(text=f"Toplam: {toplam:.2f} TL")
    
    def raporu_excele_aktar(self, treeview):
        if not treeview.get_children():
            messagebox.showwarning("Uyarı", "Aktarılacak rapor verisi yok!")
            return
        
        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyaları", "*.xlsx"), ("Tüm Dosyalar", "*.*")],
            title="Raporu Kaydet"
        )
        
        if not dosya_yolu:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Başlıkları ekle
            columns = treeview['columns']
            for col_idx, col in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=col).font = Font(bold=True)
            
            # Verileri ekle
            for row_idx, item in enumerate(treeview.get_children(), 2):
                values = treeview.item(item)['values']
                for col_idx, value in enumerate(values, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Dosyayı kaydet
            wb.save(dosya_yolu)
            messagebox.showinfo("Başarılı", "Rapor başarıyla Excel'e aktarıldı!")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel'e aktarırken hata oluştu: {str(e)}")

    def check_masa_durum(self):
        """30 dakikadır işlem yapılmayan masaları kontrol eder"""
        cursor = self.conn.cursor()
        now = datetime.now()
        threshold = now - timedelta(minutes=30)
        
        cursor.execute(
            "SELECT masa_no FROM masalar WHERE durum='Dolu' AND son_islem_tarihi < ?",
            (self.adapt_datetime(threshold),)
        )
        inactive_masalar = cursor.fetchall()
        
        for masa in inactive_masalar:
            masa_no = masa[0]
            if masa_no in self.masa_buttons:
                self.masa_buttons[masa_no].config(bg="purple")
        
        # 1 dakika sonra tekrar kontrol et
        self.root.after(60000, self.check_masa_durum)

if __name__ == "__main__":
    root = tk.Tk()
    app = CafeAdisyonProgrami(root)
    root.mainloop()
